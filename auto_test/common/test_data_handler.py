#! /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2021/3/26 21:50
# @Author : 心蓝
"""
测试数据处理
"""
import re
import json
import random

from openpyxl import load_workbook

from common import db


def get_test_data_from_excel(file, sheet_name):
    """
    获取excel文件中的用例数据
    :param file: 文件名
    :param sheet_name: 表名
    :return: [dict,dict]
    """
    # 1. 打开工作簿
    wb = load_workbook(filename=file, read_only=True)
    # 2. 获取sheet
    sh = wb[sheet_name]
    row = sh.max_row
    column = sh.max_column
    # 3. 读取数据
    data = []
    # 获取第一行拿到所有的key
    keys = []
    for i in range(1, column+1):
        keys.append(sh.cell(1, i).value)

    # 循环每一行，组成字典
    for i in range(2, row+1):
        # 循环每一行的列
        # 搞一个临时变量用来存放每一行的数据
        temp = {}
        for j in range(1, column+1):
            # 每个单元格而就是一个键值对
            # 获取对应列的键，注意列是1开头，索引是0开头
            # key = keys[j-1]
            # value = sh.cell(i, j).value
            # temp[key] = value
            temp[keys[j-1]] = sh.cell(i, j).value
        # 把request，expect_data json数据转换成python对象
        # try:
        #     temp['request'] = json.loads(temp['request'])
        #     temp['expect_data'] = json.loads(temp['expect_data'])
        # except json.decoder.JSONDecodeError:
        #     raise ValueError('用例数据json格式错误！')
        # 把每一行数据形成的字典添加到data列表中
        data.append(temp)

    return data


def generate_phone():
    """
    随机生成一个手机号码
    :return:
    """
    # 1. 1开头
    # 2. 11位
    # 3. 第二数字 3-9
    # 最好按照实际项目为准
    phone = ['158']
    # 剩下8位
    for i in range(8):
        phone.append(str(random.randint(0, 9)))
    return ''.join(phone)


def generate_no_use_phone(sql='select id from member where mobile_phone={}'):
    """
    随机生成没有使用过的手机号码
    :param sql: 校验sql模板
    :return:
    """
    while True:
        phone = generate_phone()
        sql = sql.format(phone)
        if not db.exist(sql):
            return phone


def replace_args_by_re(json_s, obj):
    """
    通过正则表达式动态的替换参数
    :param json_s:
    :param obj:
    :return:
    """
    # 1. 找出所有的槽位中的变量名
    args = re.findall('#(.+?)#', json_s)

    for arg in args:
        # 2. 找到obj中对应的属性
        value = getattr(obj, arg, None)
        if value:
            json_s = json_s.replace('#{}#'.format(arg), str(value))

    return json_s


if __name__ == '__main__':
    class Env:
        name = '心蓝'
        age = 18

    s = '{"name":"#name#", "age":#age#}'

    res = replace_args_by_re(s, Env)
    print(res)

    # res = get_test_data_from_excel('../test_data/testcases.xlsx', 'register')
    # print(res[0])
    # res = generate_phone()
    # print(res)
    # res = generate_no_use_phone()
    # print(res)